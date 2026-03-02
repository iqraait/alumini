from django.contrib import admin
from django.utils.html import format_html
from .models import AlumniRegistration, Department
import json


@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    """
    Admin configuration for Department model
    """
    list_display = (
        'name',
        'alumni_count',
        'created_at',
        'updated_at',
    )
    
    list_filter = ('created_at',)
    
    search_fields = ('name',)
    
    ordering = ('name',)
    
    list_per_page = 25
    
    # Custom method to show number of alumni in each department
    def alumni_count(self, obj):
        count = obj.alumni.count()
        url = f"/admin/myapp/alumniregistration/?iqraa_department__id__exact={obj.id}"
        
        if count > 0:
            return format_html('<a href="{}" style="font-weight: bold; color: #2563eb;">{} Alumni</a>', url, count)
        else:
            return format_html('<span style="color: #64748b;">0 Alumni</span>')
    
    alumni_count.short_description = "Associated Alumni"
    alumni_count.admin_order_field = 'alumni__count'
    
    # Actions
    actions = ['export_departments_csv']
    
    @admin.action(description="📥 Export selected departments to CSV")
    def export_departments_csv(self, request, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=departments.csv'
        
        writer = csv.writer(response)
        writer.writerow(['Department Name', 'Number of Alumni', 'Created At', 'Updated At'])
        
        for dept in queryset:
            writer.writerow([
                dept.name,
                dept.alumni.count(),
                dept.created_at.strftime('%Y-%m-%d %H:%M'),
                dept.updated_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response


@admin.register(AlumniRegistration)
class AlumniRegistrationAdmin(admin.ModelAdmin):
    """
    Admin configuration for AlumniRegistration model
    """
    
    # =========================
    # LIST DISPLAY
    # =========================
    list_display = (
        'full_name',
        'country',
        'phone_number',
        'email',
        'department_link',
        'iqraa_no',
        'approval_status',
        'created_at',
    )
    
    list_display_links = ('full_name', 'email')
    
    # =========================
    # FILTERS
    # =========================
    list_filter = (
        'country',
        'iqraa_department',
        'alumni_activities_interest',
        'hr_approved',
        'conform_register',
        'created_at',
    )
    
    # =========================
    # SEARCH FIELDS
    # =========================
    search_fields = (
        'full_name',
        'phone_number',
        'email',
        'iqraa_no',
        'iqraa_designation',
        'father_name',
        'mother_name',
        'spouse_name',
        'current_city',
        'professional_field',
    )
    
    # =========================
    # ORDERING
    # =========================
    ordering = ('-created_at',)
    list_per_page = 25
    date_hierarchy = 'created_at'
    
    # =========================
    # READONLY FIELDS
    # =========================
    readonly_fields = (
        'created_at',
        'updated_at',
        'display_children',
        'display_family_summary',
    )
    
    # =========================
    # FIELDSETS (FORM LAYOUT)
    # =========================
    fieldsets = (
        # Personal Information
        (
            '👤 PERSONAL INFORMATION',
            {
                'fields': (
                    ('full_name', 'country'),
                    ('iqraa_no', 'email'),
                    ('phone_number', 'whatsapp_number'),
                    'indian_address',
                ),
                'description': 'Basic personal and contact details'
            }
        ),
        
        # Father's Information
        (
            '👨 FATHER\'S INFORMATION',
            {
                'classes': ('wide', 'collapse'),
                'fields': (
                    'father_name',
                    ('father_iqraa_num', 'father_address'),
                ),
                'description': 'Either provide Father\'s Iqraa number OR complete address'
            }
        ),
        
        # Mother's Information
        (
            '👩 MOTHER\'S INFORMATION',
            {
                'classes': ('wide', 'collapse'),
                'fields': (
                    'mother_name',
                    ('mother_iqraa_num', 'mother_address'),
                ),
                'description': 'Either provide Mother\'s Iqraa number OR complete address'
            }
        ),
        
        # Spouse Information
        (
            '💑 SPOUSE INFORMATION',
            {
                'classes': ('wide', 'collapse'),
                'fields': (
                    'spouse_name',
                    ('spouse_iqraa_num', 'spouse_address'),
                ),
                'description': 'Fill only if applicable. Either provide Spouse\'s Iqraa number OR complete address'
            }
        ),
        
        # Children Information
        (
            '👶 CHILDREN INFORMATION',
            {
                'fields': ('display_children',),
                'classes': ('wide',),
                'description': 'Children details with Iqraa numbers or addresses'
            }
        ),
        
        # Grandparents
        (
            '👴 GRANDPARENTS',
            {
                'fields': ('grandparents',),
                'classes': ('collapse',),
            }
        ),
        
        # IQRAA Association
        (
            '🏫 IQRAA ASSOCIATION',
            {
                'fields': (
                    ('tenure_from', 'tenure_to'),
                    ('iqraa_designation', 'iqraa_department'),
                    'iqraa_campus',
                    'memorable_experience',
                ),
                'description': 'Details about time at Iqraa'
            }
        ),
        
        # Current Status
        (
            '📍 CURRENT STATUS',
            {
                'fields': (
                    ('current_city', 'professional_field'),
                    ('alumni_activities_interest', 'hear_about_us'),
                ),
            }
        ),
        
        # Family Summary (Read-only)
        (
            '📋 FAMILY SUMMARY',
            {
                'fields': ('display_family_summary',),
                'classes': ('wide', 'collapse'),
            }
        ),
        
        # Approval Workflow
        (
            '✅ APPROVAL WORKFLOW',
            {
                'fields': (
                    ('hr_approved', 'conform_register'),
                ),
                'classes': ('wide',),
                'description': 'Manage approval status'
            }
        ),
        
        # System Information
        (
            '⚙️ SYSTEM INFORMATION',
            {
                'fields': ('created_at', 'updated_at'),
                'classes': ('collapse',),
            }
        ),
    )
    
    # =========================
    # FORM WIDGETS
    # =========================
    autocomplete_fields = ['iqraa_department']  # Makes department searchable
    raw_id_fields = []  # Add if you want raw ID inputs for large tables
    filter_horizontal = []  # For many-to-many fields
    
    # =========================
    # CUSTOM METHODS
    # =========================
    
    def department_link(self, obj):
        """Display department as clickable link"""
        if obj.iqraa_department:
            url = f"/admin/myapp/department/{obj.iqraa_department.id}/change/"
            return format_html(
                '<a href="{}" style="background: #e0f2fe; padding: 4px 8px; border-radius: 20px; color: #0369a1; text-decoration: none; font-weight: 500;">{}</a>',
                url, obj.iqraa_department.name
            )
        return format_html('<span style="color: #64748b;">—</span>')
    
    department_link.short_description = "Department"
    department_link.admin_order_field = 'iqraa_department__name'
    
    def approval_status(self, obj):
        """Display approval status with colored badges"""
        if obj.hr_approved and obj.conform_register:
            return format_html(
                '<span style="background: #dcfce7; color: #166534; padding: 4px 8px; border-radius: 20px; font-weight: 500;">✓ Fully Approved</span>'
            )
        elif obj.hr_approved:
            return format_html(
                '<span style="background: #fef9c3; color: #854d0e; padding: 4px 8px; border-radius: 20px; font-weight: 500;">⏳ HR Approved</span>'
            )
        else:
            return format_html(
                '<span style="background: #fee2e2; color: #991b1b; padding: 4px 8px; border-radius: 20px; font-weight: 500;">⌛ Pending</span>'
            )
    
    approval_status.short_description = "Status"
    approval_status.admin_order_field = 'hr_approved'
    
    def display_children(self, obj):
        """Display children data in a readable format"""
        if not obj.children:
            return format_html('<span style="color: #64748b;">— No children information provided —</span>')
        
        try:
            children = obj.children if isinstance(obj.children, list) else json.loads(obj.children)
            if not children:
                return format_html('<span style="color: #64748b;">— No children information provided —</span>')
            
            html = '<div style="margin: 10px 0;">'
            for idx, child in enumerate(children, 1):
                name = child.get('name', 'Unknown')
                age = child.get('age', '')
                iqraa = child.get('iqraa', '')
                address = child.get('address', '')
                
                # Background color based on whether iqraa or address is provided
                bg_color = '#f0fdf4' if iqraa else '#fff7ed' if address else '#fef2f2'
                
                html += f'''
                <div style="background: {bg_color}; border-left: 4px solid #2563eb; padding: 12px; margin-bottom: 8px; border-radius: 0 8px 8px 0;">
                    <strong style="color: #1e40af; font-size: 1.1rem;">👶 Child #{idx}: {name}</strong>
                    {f'<div style="margin-top: 5px;"><span style="color: #4b5563;">📅 Age:</span> {age}</div>' if age else ''}
                    {f'<div style="margin-top: 5px;"><span style="color: #4b5563;">🔢 Iqraa:</span> <span style="background: #dbeafe; padding: 2px 8px; border-radius: 12px;">{iqraa}</span></div>' if iqraa else ''}
                    {f'<div style="margin-top: 5px;"><span style="color: #4b5563;">📍 Address:</span> {address}</div>' if address else ''}
                </div>
                '''
            
            html += '</div>'
            return format_html(html)
        except:
            return obj.children
    
    display_children.short_description = "Children Details"
    
    def display_family_summary(self, obj):
        """Display a summary of family information"""
        html = '<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px;">'
        
        # Father summary
        html += '<div style="background: #f8fafc; padding: 12px; border-radius: 8px;">'
        html += f'<strong style="color: #2563eb;">👨 Father:</strong> {obj.father_name}<br>'
        if obj.father_iqraa_num:
            html += f'<span style="color: #059669;">✓ Iqraa: {obj.father_iqraa_num}</span>'
        elif obj.father_address:
            html += f'<span style="color: #b45309;">📍 Address provided</span>'
        else:
            html += '<span style="color: #dc2626;">✗ No information</span>'
        html += '</div>'
        
        # Mother summary
        html += '<div style="background: #f8fafc; padding: 12px; border-radius: 8px;">'
        html += f'<strong style="color: #2563eb;">👩 Mother:</strong> {obj.mother_name}<br>'
        if obj.mother_iqraa_num:
            html += f'<span style="color: #059669;">✓ Iqraa: {obj.mother_iqraa_num}</span>'
        elif obj.mother_address:
            html += f'<span style="color: #b45309;">📍 Address provided</span>'
        else:
            html += '<span style="color: #dc2626;">✗ No information</span>'
        html += '</div>'
        
        # Spouse summary (if exists)
        if obj.spouse_name:
            html += '<div style="background: #f8fafc; padding: 12px; border-radius: 8px;">'
            html += f'<strong style="color: #2563eb;">💑 Spouse:</strong> {obj.spouse_name}<br>'
            if obj.spouse_iqraa_num:
                html += f'<span style="color: #059669;">✓ Iqraa: {obj.spouse_iqraa_num}</span>'
            elif obj.spouse_address:
                html += f'<span style="color: #b45309;">📍 Address provided</span>'
            else:
                html += '<span style="color: #dc2626;">✗ No information</span>'
            html += '</div>'
        
        # Children count
        children_count = len(obj.children) if obj.children else 0
        html += f'<div style="background: #f8fafc; padding: 12px; border-radius: 8px;">'
        html += f'<strong style="color: #2563eb;">👶 Children:</strong> {children_count} child(ren)'
        html += '</div>'
        
        html += '</div>'
        return format_html(html)
    
    display_family_summary.short_description = "Family Summary"
    
    # =========================
    # ACTIONS
    # =========================
    actions = [
        'mark_hr_approved',
        'mark_hr_pending',
        'mark_conformed',
        'mark_not_conformed',
        'export_as_csv',
        'export_as_excel',
    ]
    
    @admin.action(description="✅ Mark selected as HR Approved")
    def mark_hr_approved(self, request, queryset):
        updated = queryset.update(hr_approved=True)
        self.message_user(
            request,
            f"✓ {updated} alumni marked as HR Approved.",
            level='SUCCESS'
        )
    
    @admin.action(description="⏳ Mark selected as HR Pending")
    def mark_hr_pending(self, request, queryset):
        updated = queryset.update(hr_approved=False)
        self.message_user(
            request,
            f"⏱️ {updated} alumni marked as HR Pending.",
            level='WARNING'
        )
    
    @admin.action(description="✅ Mark selected as Conformed")
    def mark_conformed(self, request, queryset):
        updated = queryset.update(conform_register=True)
        self.message_user(
            request,
            f"✓ {updated} alumni marked as Conformed.",
            level='SUCCESS'
        )
    
    @admin.action(description="❌ Mark selected as Not Conformed")
    def mark_not_conformed(self, request, queryset):
        updated = queryset.update(conform_register=False)
        self.message_user(
            request,
            f"✗ {updated} alumni marked as Not Conformed.",
            level='WARNING'
        )
    
    @admin.action(description="📥 Export selected as CSV")
    def export_as_csv(self, request, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=alumni_export_{request.user.username}.csv'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = [
            'Full Name', 'Country', 'Iqraa No', 'Phone', 'WhatsApp', 'Email',
            'Father Name', 'Father Iqraa', 'Father Address',
            'Mother Name', 'Mother Iqraa', 'Mother Address',
            'Spouse Name', 'Spouse Iqraa', 'Spouse Address',
            'Children Count', 'Tenure From', 'Tenure To',
            'Department', 'Designation', 'Campus',
            'Current City', 'Professional Field', 'Activities Interest',
            'HR Approved', 'Conformed', 'Created At'
        ]
        writer.writerow(headers)
        
        # Write data
        for obj in queryset:
            writer.writerow([
                obj.full_name,
                obj.country,
                obj.iqraa_no,
                obj.phone_number,
                obj.whatsapp_number,
                obj.email,
                obj.father_name,
                obj.father_iqraa_num,
                obj.father_address,
                obj.mother_name,
                obj.mother_iqraa_num,
                obj.mother_address,
                obj.spouse_name,
                obj.spouse_iqraa_num,
                obj.spouse_address,
                len(obj.children) if obj.children else 0,
                obj.tenure_from,
                obj.tenure_to,
                obj.iqraa_department.name if obj.iqraa_department else '',
                obj.iqraa_designation,
                obj.iqraa_campus,
                obj.current_city,
                obj.professional_field,
                obj.alumni_activities_interest,
                'Yes' if obj.hr_approved else 'No',
                'Yes' if obj.conform_register else 'No',
                obj.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        
        return response
    
    @admin.action(description="📊 Export selected as Excel")
    def export_as_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alumni Export"
        
        # Write headers
        headers = [
            'Full Name', 'Country', 'Iqraa No', 'Phone', 'WhatsApp', 'Email',
            'Father Name', 'Father Iqraa', 'Father Address',
            'Mother Name', 'Mother Iqraa', 'Mother Address',
            'Spouse Name', 'Spouse Iqraa', 'Spouse Address',
            'Children Count', 'Tenure From', 'Tenure To',
            'Department', 'Designation', 'Campus',
            'Current City', 'Professional Field', 'Activities Interest',
            'HR Approved', 'Conformed', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Write data
        for row, obj in enumerate(queryset, 2):
            data = [
                obj.full_name,
                obj.country,
                obj.iqraa_no,
                obj.phone_number,
                obj.whatsapp_number,
                obj.email,
                obj.father_name,
                obj.father_iqraa_num,
                obj.father_address,
                obj.mother_name,
                obj.mother_iqraa_num,
                obj.mother_address,
                obj.spouse_name,
                obj.spouse_iqraa_num,
                obj.spouse_address,
                len(obj.children) if obj.children else 0,
                obj.tenure_from,
                obj.tenure_to,
                obj.iqraa_department.name if obj.iqraa_department else '',
                obj.iqraa_designation,
                obj.iqraa_campus,
                obj.current_city,
                obj.professional_field,
                obj.alumni_activities_interest,
                'Yes' if obj.hr_approved else 'No',
                'Yes' if obj.conform_register else 'No',
                obj.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=alumni_export_{request.user.username}.xlsx'
        
        wb.save(response)
        return response
    
    # =========================
    # OVERRIDE QUERYSET
    # =========================
    def get_queryset(self, request):
        """Optimize queryset with select_related"""
        return super().get_queryset(request).select_related('iqraa_department')
    
    # =========================
    # SAVE RELATED
    # =========================
    save_on_top = True
    save_as = True  # Allow save as new
    show_full_result_count = False


# Optional: Inline for Department to show related alumni
class AlumniInline(admin.TabularInline):
    model = AlumniRegistration
    fields = ['full_name', 'email', 'phone_number', 'created_at']
    readonly_fields = ['full_name', 'email', 'phone_number', 'created_at']
    extra = 0
    can_delete = False
    show_change_link = True
    
    def has_add_permission(self, request, obj=None):
        return False


# Uncomment if you want to see alumni inline in Department admin
# DepartmentAdmin.inlines = [AlumniInline]